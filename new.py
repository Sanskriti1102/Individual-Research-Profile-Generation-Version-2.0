import streamlit as st
import openpyxl
import logging
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import tempfile
import os

# --- Your scraper class (same as before) ---
class GoogleScholarScraper:
    def __init__(self, professor_name, scholar_id):
        self.professor_name = professor_name
        self.scholar_id = scholar_id
        self.driver = None
        self.base_url = f'https://scholar.google.com/citations?hl=en&user={self.scholar_id}&view_op=list_works&sortby=pubdate'
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.wait = None
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    def setup_driver(self):
        options = ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        self.driver = Chrome(options=options)
        self.driver.maximize_window()
        self.wait = WebDriverWait(self.driver, 10)

    def teardown_driver(self):
        if self.driver:
            self.driver.quit()

    def scroll_to_load_all(self):
        count = 0
        while True:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            try:
                more_button = self.wait.until(EC.element_to_be_clickable((By.ID, 'gsc_bpf_more')))
                more_button.click()
                self.wait.until(EC.staleness_of(more_button))
            except Exception:
                break
            new_count = len(self.driver.find_elements(By.CSS_SELECTOR, "tr.gsc_a_tr"))
            if new_count == count:
                break
            count = new_count

    def scrape_publications(self):
        self.sheet.append(['Title', 'Authors', 'Publication Date', 'Document Type', 'Link'])
        rows = self.driver.find_elements(By.CSS_SELECTOR, "tr.gsc_a_tr")
        for idx, row in enumerate(rows, 1):
            try:
                title_el = row.find_element(By.CSS_SELECTOR, "td.gsc_a_t a")
                title = title_el.text.strip()
                link = title_el.get_attribute("href")
                author_journal = row.find_elements(By.CSS_SELECTOR, "td.gsc_a_t div.gs_gray")
                authors = author_journal[0].text.strip() if len(author_journal) > 0 else ""
                pub_date = row.find_element(By.CSS_SELECTOR, "td.gsc_a_y span").text.strip()
                self.sheet.append([title, authors, pub_date, 'Other', link])
            except Exception as e:
                logging.error(f"Error scraping publication {idx}: {e}")

    def save_to_excel(self, filepath):
        self.workbook.save(filepath)

    def run(self, filepath):
        self.setup_driver()
        self.driver.get(self.base_url)
        self.scroll_to_load_all()
        self.scrape_publications()
        self.teardown_driver()
        self.save_to_excel(filepath)


def fetch_scholar_id(professor_name, excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    scholar_id = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, s_id, _ = row
        if name and name.strip().lower() == professor_name.strip().lower():
            scholar_id = s_id
            break
    wb.close()
    return scholar_id


# --- Streamlit UI ---
st.title("Google Scholar Profile Scraper")

uploaded_file = st.file_uploader("Upload Excel with Professors (Name, Scholar ID, ...)", type=["xlsx"])
professor_name = st.text_input("Enter Professor's Full Name")

if st.button("Scrape Publications"):
    if not uploaded_file:
        st.error("Please upload the Excel file containing professor names and Scholar IDs.")
    elif not professor_name.strip():
        st.error("Please enter a valid professor's name.")
    else:
        with st.spinner('Fetching Scholar ID and scraping... This may take a while.'):
            # Save uploaded Excel temporarily to disk
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded_file.read())
                tmp_path = tmp.name

            scholar_id = fetch_scholar_id(professor_name, tmp_path)
            if not scholar_id:
                st.error(f"Professor '{professor_name}' not found in the uploaded Excel.")
                os.unlink(tmp_path)  # delete temp file
            else:
                try:
                    # Save output Excel to temp file
                    output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
                    scraper = GoogleScholarScraper(professor_name, scholar_id)
                    scraper.run(output_path)

                    # Show success + download button
                    st.success(f"Scraping completed for {professor_name}!")
                    with open(output_path, "rb") as f:
                        st.download_button(
                            label="Download Publications Excel",
                            data=f,
                            file_name=f"{professor_name}_publications.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    os.unlink(output_path)
                    os.unlink(tmp_path)
                except Exception as e:
                    st.error(f"An error occurred during scraping: {e}")
                    os.unlink(tmp_path)
